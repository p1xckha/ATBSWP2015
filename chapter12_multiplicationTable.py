# -*- coding: utf-8 -*-
"""
Automate The Boring Stuff With Python 
chapter 12
Practice Projects: Multiplication Table Maker

if n=4,
the multiplication table is output in the format of excel file as follows:

Excel file
    
   1  2  3  4
1  1  2  3  4
2  2  4  6  8
3  3  6  9 12
4  4  8 12 16

"""

import openpyxl
from openpyxl.styles import Font

import sys

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Multiplication Table"

n = 20 # defalt value
if len(sys.argv) > 1:
    n = int(sys.argv[1])

font = Font(bold=True)

# labels in A2, A3,...
for row in range(2, n+2):
    sheet.cell(row=row, column=1).value = row - 1
    sheet.cell(row=row, column=1).font = font

# labels in B1, C1,...
for column in range(2, n+2):
    sheet.cell(row=1, column=column).value = column - 1
    sheet.cell(row=1, column=column).font = font

# multiplication
for row in range(2, n+2):
    for column in range(2, n+2):
        sheet.cell(row=row, column=column).value = (row-1) * (column-1)

sheet.freeze_panes = 'B2' # 1st row and column are fixed
wb.save("multiplicationTable.xlsx")

